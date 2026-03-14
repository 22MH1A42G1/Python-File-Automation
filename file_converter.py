"""
file_converter.py
-----------------
Convert files between common formats.

Supported conversions
---------------------
    csv-to-json   Convert a CSV file to a JSON file.
    json-to-csv   Convert a JSON file (array of objects) to a CSV file.
    excel-to-csv  Convert each sheet of an Excel workbook to a separate CSV file.

Usage
-----
    python file_converter.py csv-to-json  <input.csv>  [output.json]
    python file_converter.py json-to-csv  <input.json> [output.csv]
    python file_converter.py excel-to-csv <input.xlsx> [output_dir]

Examples
--------
    python file_converter.py csv-to-json  data/sales.csv  data/sales.json
    python file_converter.py json-to-csv  data/sales.json data/sales.csv
    python file_converter.py excel-to-csv report.xlsx    ./csv_output/
"""

import argparse
import csv
import json
import os
import sys


# ---------------------------------------------------------------------------
# CSV → JSON
# ---------------------------------------------------------------------------

def csv_to_json(input_path: str, output_path: str | None = None) -> str:
    """Convert a CSV file to a JSON file and return the output path."""
    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path!r}")

    if output_path is None:
        base, _ = os.path.splitext(input_path)
        output_path = base + ".json"

    with open(input_path, newline="", encoding="utf-8") as csv_file:
        reader = csv.DictReader(csv_file)
        rows = list(reader)

    with open(output_path, "w", encoding="utf-8") as json_file:
        json.dump(rows, json_file, indent=4, ensure_ascii=False)

    print(f"CSV → JSON: {input_path!r}  →  {output_path!r}  ({len(rows)} rows)")
    return output_path


# ---------------------------------------------------------------------------
# JSON → CSV
# ---------------------------------------------------------------------------

def json_to_csv(input_path: str, output_path: str | None = None) -> str:
    """Convert a JSON file (array of objects) to a CSV file and return the output path."""
    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path!r}")

    if output_path is None:
        base, _ = os.path.splitext(input_path)
        output_path = base + ".csv"

    with open(input_path, encoding="utf-8") as json_file:
        data = json.load(json_file)

    if not isinstance(data, list):
        raise ValueError("JSON file must contain a top-level array of objects.")
    if not data:
        raise ValueError("JSON array is empty — nothing to convert.")

    fieldnames = list(data[0].keys())

    with open(output_path, "w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)

    print(f"JSON → CSV: {input_path!r}  →  {output_path!r}  ({len(data)} rows)")
    return output_path


# ---------------------------------------------------------------------------
# Excel → CSV
# ---------------------------------------------------------------------------

def excel_to_csv(input_path: str, output_dir: str | None = None) -> list[str]:
    """Convert each sheet of an Excel workbook to a CSV file.

    Returns a list of the output CSV file paths.
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit(
            "openpyxl is required for Excel conversion.\n"
            "Install it with:  pip install openpyxl"
        )

    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path!r}")

    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(input_path))

    os.makedirs(output_dir, exist_ok=True)

    workbook = openpyxl.load_workbook(input_path, data_only=True)
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    output_paths: list[str] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        safe_sheet = sheet_name.replace(" ", "_")
        out_file = os.path.join(output_dir, f"{base_name}_{safe_sheet}.csv")

        with open(out_file, "w", newline="", encoding="utf-8") as csv_file:
            writer = csv.writer(csv_file)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        row_count = sheet.max_row
        print(f"Excel → CSV: sheet {sheet_name!r}  →  {out_file!r}  ({row_count} rows)")
        output_paths.append(out_file)

    workbook.close()
    print(f"\nTotal sheets converted: {len(output_paths)}")
    return output_paths


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert files between CSV, JSON, and Excel formats.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # csv-to-json
    p_c2j = sub.add_parser("csv-to-json", help="Convert CSV → JSON.")
    p_c2j.add_argument("input", help="Path to the input CSV file.")
    p_c2j.add_argument("output", nargs="?", help="Path for the output JSON file (optional).")

    # json-to-csv
    p_j2c = sub.add_parser("json-to-csv", help="Convert JSON → CSV.")
    p_j2c.add_argument("input", help="Path to the input JSON file.")
    p_j2c.add_argument("output", nargs="?", help="Path for the output CSV file (optional).")

    # excel-to-csv
    p_e2c = sub.add_parser("excel-to-csv", help="Convert Excel → CSV (one file per sheet).")
    p_e2c.add_argument("input", help="Path to the input Excel file (.xlsx).")
    p_e2c.add_argument("output_dir", nargs="?", help="Directory for output CSV files (optional).")

    return parser


if __name__ == "__main__":
    args = _build_parser().parse_args()

    if args.command == "csv-to-json":
        csv_to_json(args.input, args.output)
    elif args.command == "json-to-csv":
        json_to_csv(args.input, args.output)
    elif args.command == "excel-to-csv":
        excel_to_csv(args.input, args.output_dir)
